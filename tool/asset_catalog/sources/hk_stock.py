"""Hong Kong securities catalog adapter (HKEX official list).

Source
------
HKEX publishes a single Excel workbook containing every listed
security: equities, ETFs, REITs, derivative warrants, CBBCs, etc.

    https://www.hkex.com.hk/eng/services/trading/securities/securitieslists/ListOfSecurities.xlsx

We keep equities / ETFs / REITs / DWs (derivative warrants tracked by
many investors), and drop CBBCs and inline warrants — they have very
short lifecycles and would dominate the row count without adding search
value. Names are normalised through OpenCC so traditional and
mixed-script HKEX entries become canonical simplified Chinese before
``build.py`` runs pinyin generation.

Sheet structure (HKEX format, may evolve):

    row 0..2: title block / "as of" stamp
    row 3:    header row with columns
              ``Stock Code`` | ``Name of Securities`` |
              ``Category`` (EQUITY / ETF / REIT / DW / CBBC / ...) |
              ``Sub-Category`` | ``Board Lot`` | ``CCASS`` | ...
    row 4..N: data rows

The Stock Code is a numeric string padded to 4-5 digits; we left-pad to
4 digits and append ``.HK`` so it matches Yahoo / Sina symbology already
used elsewhere in NaviWealth.
"""
from __future__ import annotations

import argparse
import logging
import pathlib
import re
import sys
from collections.abc import Callable, Mapping
from typing import Any

from tool.asset_catalog.common import make_row, to_simplified, write_rows

LOG = logging.getLogger("asset_catalog.hk_stock")

MIN_HK_STOCK_ROWS = 1_500
YFINANCE_PAGE_SIZE = 250
YFINANCE_MAX_RESULTS = 20_000

HKEX_URL = (
    "https://www.hkex.com.hk/eng/services/trading/securities/"
    "securitieslists/ListOfSecurities.xlsx"
)

CATEGORY_TYPE_MAP = {
    "equity": "stock",
    "equities": "stock",
    "etf": "etf",
    "etp": "etf",  # exchange-traded products (legacy label)
    "reit": "stock",  # treated as a stock for catalog purposes
    "dw": "stock",  # derivative warrants — keep, but flag via aliases
}

# Categories we explicitly drop. Anything not in TYPE_MAP and not in
# DROPPED_CATEGORIES is logged and skipped to make schema drift visible.
DROPPED_CATEGORIES = {
    "cbbc",
    "inline warrant",
    "iw",
    "structured product",
}


def _normalise_category(value: str) -> str:
    return (value or "").strip().lower()


def parse_hkex_row(
    code: object,
    name: object,
    category: object,
) -> dict[str, str] | None:
    """Translate one HKEX row to a CSV row, or ``None`` to drop.

    Inputs are deliberately ``object``-typed because openpyxl returns a
    mix of ``int`` / ``str`` / ``None`` cells.
    """
    code_str = str(code).strip() if code is not None else ""
    name_str = str(name).strip() if name is not None else ""
    if not code_str or not name_str:
        return None
    if not code_str.isdigit():
        return None

    cat = _normalise_category(str(category) if category is not None else "")
    asset_type = CATEGORY_TYPE_MAP.get(cat)
    if asset_type is None:
        if cat in DROPPED_CATEGORIES:
            return None
        # Unknown category — drop silently in production, but emit a
        # debug log so build runs surface upstream changes.
        LOG.debug("hkex: unknown category %r for %s (%s)", cat, code_str, name_str)
        return None

    # HKEX currently emits five-character strings such as ``00700`` while
    # Yahoo/Sina use ``0700.HK``. Removing leading zeroes before padding keeps
    # ordinary listings canonical while preserving real five-digit products.
    padded = str(int(code_str)).zfill(4)
    symbol = f"{padded}.HK"
    name_cn = to_simplified(name_str)
    return make_row(
        symbol=symbol,
        market="hk_stock",
        type_name=asset_type,
        currency="HKD",
        name_en=name_str if _looks_latin(name_str) else "",
        name_cn=name_cn if not _looks_latin(name_cn) else "",
    )


def _looks_latin(text: str) -> bool:
    """Heuristic — most HKEX names are pure Chinese OR pure English; very
    few mix. We use a >50% ASCII-letters ratio to decide which slot to
    populate."""
    if not text:
        return False
    letters = sum(1 for c in text if c.isalpha())
    if letters == 0:
        return False
    ascii_letters = sum(1 for c in text if c.isascii() and c.isalpha())
    return ascii_letters / letters > 0.5


# ---------------------------------------------------------------------------
# Network path
# ---------------------------------------------------------------------------


class FetchError(RuntimeError):
    pass


def fetch_hkex(url: str = HKEX_URL, timeout: int = 60) -> bytes:
    try:
        import requests  # type: ignore
    except ImportError as exc:
        raise FetchError("requests not installed; run pip install -r requirements.txt") from exc
    headers = {"User-Agent": "NaviWealth-CatalogBuilder/1.0"}
    try:
        resp = requests.get(url, timeout=timeout, headers=headers)
    except requests.RequestException as exc:
        raise FetchError(f"HKEX request failed for {url}: {exc}") from exc
    if resp.status_code != 200:
        raise FetchError(f"HKEX HTTP {resp.status_code} for {url}")
    return resp.content


def parse_workbook(blob: bytes) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise FetchError("openpyxl not installed; run pip install -r requirements.txt") from exc
    import io

    try:
        wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    except Exception as exc:
        raise FetchError(f"HKEX: invalid workbook: {exc}") from exc
    sheet = wb.worksheets[0]

    rows = sheet.iter_rows(values_only=True)
    headers: list[str] | None = None
    out: list[dict[str, str]] = []
    for row in rows:
        if not row:
            continue
        if headers is None:
            normalised = [str(c).strip().lower() if c is not None else "" for c in row]
            if "stock code" in normalised and any("name" in c for c in normalised):
                headers = normalised
            continue
        try:
            code_idx = headers.index("stock code")
        except ValueError:
            raise FetchError("HKEX: 'Stock Code' column not found")
        name_idx = next(
            (i for i, h in enumerate(headers) if "name of securit" in h or h == "name"),
            -1,
        )
        if name_idx < 0:
            raise FetchError("HKEX: name column not found")
        cat_idx = next(
            (i for i, h in enumerate(headers) if "category" in h and "sub" not in h),
            -1,
        )
        parsed = parse_hkex_row(
            code=row[code_idx] if code_idx < len(row) else None,
            name=row[name_idx] if name_idx < len(row) else None,
            category=row[cat_idx] if 0 <= cat_idx < len(row) else None,
        )
        if parsed:
            out.append(parsed)
    LOG.info("hkex: %d catalog rows", len(out))
    return out


def parse_yfinance_quote(
    quote: Mapping[str, Any],
    *,
    type_name: str,
) -> dict[str, str] | None:
    """Translate one Yahoo screener quote into a catalog row.

    Yahoo reports derivative warrants, CBBCs, and ordinary shares alike as
    ``EQUITY``. The fallback therefore keeps only the 0001-9999 ordinary/GEM
    code space for equities. ETFs are obtained from Yahoo's separate ETF
    screener.
    """
    expected_quote_type = "ETF" if type_name == "etf" else "EQUITY"
    if str(quote.get("quoteType") or "").upper() != expected_quote_type:
        return None
    if str(quote.get("exchange") or "").upper() != "HKG":
        return None
    if str(quote.get("currency") or "").upper() != "HKD":
        return None

    match = re.fullmatch(r"(\d{4,5})\.HK", str(quote.get("symbol") or "").upper())
    if match is None:
        return None
    code = int(match.group(1))
    if code <= 0 or code > 9_999:
        return None

    name = str(quote.get("shortName") or quote.get("longName") or "").strip()
    if not name:
        return None

    symbol = f"{code:04d}.HK"
    name_cn = to_simplified(name)
    return make_row(
        symbol=symbol,
        market="hk_stock",
        type_name=type_name,
        currency="HKD",
        name_en=name if _looks_latin(name) else "",
        name_cn=name_cn if not _looks_latin(name_cn) else "",
    )


def _fetch_yfinance_screen(
    screen: Callable[..., Mapping[str, Any]],
    query: object,
) -> list[Mapping[str, Any]]:
    """Fetch every page for one Yahoo screener query."""

    def fetch_page(offset: int) -> Mapping[str, Any]:
        result = screen(
            query,
            offset=offset,
            size=YFINANCE_PAGE_SIZE,
            sortField="ticker",
            sortAsc=True,
        )
        if not isinstance(result, Mapping):
            raise FetchError(
                f"yfinance screener returned {type(result)!r}, expected a mapping"
            )
        return result

    first = fetch_page(0)
    total = first.get("total")
    if not isinstance(total, int) or total < 0 or total > YFINANCE_MAX_RESULTS:
        raise FetchError(f"yfinance screener returned invalid total: {total!r}")

    first_quotes = first.get("quotes")
    if not isinstance(first_quotes, list):
        raise FetchError("yfinance screener response has no quotes list")
    quotes: list[Mapping[str, Any]] = [
        quote for quote in first_quotes if isinstance(quote, Mapping)
    ]

    for offset in range(YFINANCE_PAGE_SIZE, total, YFINANCE_PAGE_SIZE):
        page = fetch_page(offset)
        page_quotes = page.get("quotes")
        if not isinstance(page_quotes, list):
            raise FetchError(
                f"yfinance screener page at offset {offset} has no quotes list"
            )
        if not page_quotes:
            raise FetchError(
                f"yfinance screener stopped at offset {offset} before total {total}"
            )
        quotes.extend(quote for quote in page_quotes if isinstance(quote, Mapping))

    if len(quotes) != total:
        raise FetchError(
            f"yfinance screener returned {len(quotes)} quotes, expected {total}"
        )
    symbols = [quote.get("symbol") for quote in quotes]
    if any(not isinstance(symbol, str) or not symbol for symbol in symbols):
        raise FetchError("yfinance screener returned a quote without a symbol")
    if len(set(symbols)) != len(symbols):
        raise FetchError("yfinance screener returned duplicate symbols across pages")
    return quotes


def fetch_yfinance() -> list[dict[str, str]]:
    """Fetch a conservative HK stock + ETF universe from Yahoo Finance."""
    try:
        import yfinance as yf  # type: ignore
    except ImportError as exc:
        raise FetchError(
            "yfinance not installed; run pip install -r requirements.txt"
        ) from exc

    try:
        equity_query = yf.EquityQuery(
            "and",
            [
                yf.EquityQuery("eq", ["region", "hk"]),
                yf.EquityQuery("eq", ["exchange", "HKG"]),
            ],
        )
        etf_query = yf.ETFQuery(
            "and",
            [
                yf.ETFQuery("eq", ["region", "hk"]),
                yf.ETFQuery("eq", ["exchange", "HKG"]),
            ],
        )
        equity_quotes = _fetch_yfinance_screen(yf.screen, equity_query)
        etf_quotes = _fetch_yfinance_screen(yf.screen, etf_query)
    except FetchError:
        raise
    except Exception as exc:
        raise FetchError(f"yfinance HK screener failed: {exc}") from exc

    by_symbol: dict[str, dict[str, str]] = {}
    for quote in equity_quotes:
        parsed = parse_yfinance_quote(quote, type_name="stock")
        if parsed is not None:
            by_symbol[parsed["symbol"]] = parsed
    for quote in etf_quotes:
        parsed = parse_yfinance_quote(quote, type_name="etf")
        if parsed is not None:
            # Prefer the more specific ETF classification if Yahoo returns a
            # symbol from both screeners.
            by_symbol[parsed["symbol"]] = parsed

    rows = [by_symbol[symbol] for symbol in sorted(by_symbol)]
    LOG.info(
        "yfinance: %d catalog rows (equity quotes=%d, ETF quotes=%d)",
        len(rows),
        len(equity_quotes),
        len(etf_quotes),
    )
    return rows


def fetch_hk_stock() -> list[dict[str, str]]:
    try:
        rows = parse_workbook(fetch_hkex())
    except FetchError as exc:
        LOG.warning("HKEX unavailable (%s); falling back to yfinance", exc)
    else:
        if len(rows) >= MIN_HK_STOCK_ROWS:
            return rows
        LOG.warning(
            "HKEX returned only %d rows (minimum %d); falling back to yfinance",
            len(rows),
            MIN_HK_STOCK_ROWS,
        )

    rows = fetch_yfinance()
    if len(rows) < MIN_HK_STOCK_ROWS:
        raise FetchError(
            f"yfinance returned only {len(rows)} HK rows "
            f"(minimum {MIN_HK_STOCK_ROWS})"
        )
    return rows


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output",
        type=pathlib.Path,
        default=pathlib.Path("tool/asset_catalog/sources/hk_stock.csv"),
        help="output CSV path (use '-' for stdout)",
    )
    parser.add_argument("-v", "--verbose", action="store_true")
    args = parser.parse_args(argv)
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    )
    rows = fetch_hk_stock()
    written = write_rows(rows, args.output)
    LOG.info("wrote %d rows to %s", written, args.output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
